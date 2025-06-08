from datetime import datetime
from aiogram.types import Message
from aiogram.fsm.context import FSMContext
from aiogram.types import FSInputFile
import openpyxl
from openpyxl.reader.excel import load_workbook


async def generate_excel_file(data: dict, user_id: int):
    installation = data.get('installation')
    frame = data.get('frame')
    recycling = data.get('recycling')
    type_ahu = data.get('type_ahu')
    recuperation = data.get('recuperation')
    type_glycol = data.get('type_glycol')
    percent_glycol = data.get('percent_glycol')
    percent_recycling = data.get('percent_recycling')
    temp_recycling = data.get('temp_recycling')
    winter_outside_temp = data.get('winter_outside_temp')
    winter_outside_humidity = data.get('winter_outside_humidity')
    summer_outside_temp = data.get('summer_outside_temp')
    summer_outside_humidity = data.get('summer_outside_humidity')
    winter_exhaust_temp = data.get('winter_exhaust_temp')
    winter_exhaust_humidity = data.get('winter_exhaust_humidity')
    summer_exhaust_temp = data.get('summer_exhaust_temp')
    summer_exhaust_humidity = data.get('summer_exhaust_humidity')
    winter_internal1_temp = data.get('winter_internal1_temp')
    winter_internal2_temp = data.get('winter_internal2_temp')
    summer_internal1_temp = data.get('summer_internal1_temp')
    summer_internal2_temp = data.get('summer_internal2_temp')
    supply_air = data.get('supply_air')
    exhaust_air = data.get('exhaust_air')
    supply_pressure = data.get('supply_pressure')
    exhaust_pressure = data.get('exhaust_pressure')
    valve = data.get('valve')
    supply_silence = data.get('supply_silence')
    exhaust_silence = data.get('exhaust_silence')
    supply_first_filter = data.get('supply_first_filter')
    exhaust_first_filter = data.get('exhaust_first_filter')
    supply_second_filter = data.get('supply_second_filter')
    exhaust_second_filter = data.get('exhaust_second_filter')
    confirm_supply_second_filter = data.get('confirm_supply_second_filter')
    confirm_exhaust_second_filter = data.get('confirm_exhaust_second_filter')
    automation = data.get('automation')
    heating1_type = data.get('heating1_type')
    heating2_type = data.get('heating2_type')
    heating1_param = data.get('heating1_param')
    heating2_param = data.get('heating2_param')
    heating1_percent_glycol = data.get('heating1_percent_glycol')
    heating2_percent_glycol = data.get('heating2_percent_glycol')
    cooling1_type = data.get('cooling1_type')
    cooling2_type = data.get('cooling2_type')
    confirm_heating2 = data.get('confirm_heating2')
    confirm_cooling2 = data.get('confirm_cooling2')
    cooling1_param = data.get('cooling1_param')
    cooling2_param = data.get('cooling2_param')
    cooling1_glycol = data.get('cooling1_glycol')
    cooling2_glycol = data.get('cooling2_glycol')
    cooling1_percent_glycol = data.get('cooling1_percent_glycol')
    cooling2_percent_glycol = data.get('cooling2_percent_glycol')
    freon1 = data.get('freon1')
    freon2 = data.get('freon2')
    notes = data.get('notes')
    contact = data.get('contact_name')
    phone = data.get('contact_phone')
    organization = data.get('organization')

    wb = openpyxl.load_workbook("templates/ahu_template.xlsx")
    ws = wb.active

    ws['A14'] = '+' if installation == 'Внутрішня' else None
    ws['A15'] = '+' if installation == 'Наружна' else None

    ws['E14'] = '+' if frame == 'На рамі' else None
    ws['E15'] = '+' if frame == 'Без рами' else None

    ws['A47'] = '+' if recycling == 'Так' else None

    ws['I14'] = '+' if type_ahu == 'Припливна' else None
    ws['I15'] = '+' if type_ahu == 'Витяжна' else None
    ws['I16'] = '+' if type_ahu == 'Припливно-витяжна' else None

    ws['A49'] = '+' if recuperation == 'Гліколевий' else None
    ws['A51'] = '+' if recuperation == 'Пластинчатий' else None
    ws['A53'] = '+' if recuperation == 'Роторний' else None
    ws['A55'] = '+' if recuperation == 'Без рекуператора' else None

    ws['F49'] = type_glycol
    ws['K49'] = percent_glycol

    ws['F47'] = percent_recycling
    ws['K47'] = temp_recycling

    ws['F58'] = winter_outside_temp
    ws['H58'] = winter_outside_humidity
    ws['K58'] = summer_outside_temp
    ws['M58'] = summer_outside_humidity
    ws['F59'] = winter_exhaust_temp
    ws['H59'] = winter_exhaust_humidity
    ws['K59'] = summer_exhaust_temp
    ws['M59'] = summer_exhaust_humidity

    ws['F29'] = winter_internal1_temp
    ws['F33'] = winter_internal2_temp
    ws['F37'] = summer_internal1_temp
    ws['F42'] = summer_internal2_temp

    ws['G20'] = supply_air if type_ahu in ['Припливна', 'Припливно-витяжна'] else None
    ws['K20'] = exhaust_air if type_ahu in ['Витяжна', 'Припливно-витяжна'] else None
    ws['G21'] = supply_pressure if type_ahu in ['Припливна', 'Припливно-витяжна'] else None
    ws['K21'] = exhaust_pressure if type_ahu in ['Витяжна', 'Припливно-витяжна'] else None

    ws['G23'] = '+' if valve == 'Так' else None
    ws['I23'] = '+' if valve == 'Ні' else None
    ws['G62'] = supply_silence if (supply_silence in ['910', '1090', '1390', '1600']
                                   and type_ahu in ['Припливна', 'Припливно-витяжна']) else None
    ws['N62'] = exhaust_silence if (supply_silence in ['910', '1090', '1390', '1600']
                                   and type_ahu in ['Витяжна', 'Припливно-витяжна']) else None
    ws['G26'] = supply_first_filter if (supply_first_filter in ['Панельний G4', 'Кишеньковий G4', 'M5', 'F7', 'F8', 'F9']
                                    and type_ahu in ['Припливна', 'Припливно-витяжна']) else None
    ws['M26'] = exhaust_first_filter if (exhaust_first_filter in ['Панельний G4', 'Кишеньковий G4', 'M5', 'F7', 'F8', 'F9']
                                    and type_ahu in ['Витяжна', 'Припливно-витяжна']) else None
    ws['G27'] = supply_second_filter if (confirm_supply_second_filter == 'Так' and
                supply_second_filter in ['Панельний G4', 'Кишеньковий G4', 'M5', 'F7', 'F8', 'F9']
                and type_ahu in ['Припливна', 'Припливно-витяжна']) else None
    ws['M27'] = exhaust_second_filter if (confirm_exhaust_second_filter == 'Так' and
                                          exhaust_second_filter in ['Панельний G4', 'Кишеньковий G4', 'M5', 'F7', 'F8',
                                                                   'F9']
                                          and type_ahu in ['Витяжна', 'Припливно-витяжна']) else None

    ws['E65'] = '+' if automation == 'Так' else None
    ws['E66'] = '+' if automation == 'Ні' else None

    ws['N29'] = '+' if heating1_type == 'Водяний' and type_ahu in ['Припливна', 'Припливно-витяжна'] else None
    ws['N30'] = '+' if heating1_type == 'Електричний' and type_ahu in ['Припливна', 'Припливно-витяжна'] else None
    ws['N33'] = '+' if (heating2_type == 'Водяний' and type_ahu in ['Припливна', 'Припливно-витяжна']
                        and confirm_heating2 == 'Так') else None
    ws['N34'] = '+' if (heating2_type == 'Електричний' and type_ahu in ['Припливна', 'Припливно-витяжна']
                        and confirm_heating2 == 'Так') else None

    ws['F30'] = heating1_param if (heating1_type == 'Водяний'
                                   and type_ahu in ['Припливна', 'Припливно-витяжна']) else None
    ws['F34'] = heating2_param if (heating2_type == 'Водяний'
                                   and confirm_heating2 == 'Так'
                                   and type_ahu in ['Припливна','Припливно-витяжна']) else None
    ws['F31'] = heating1_percent_glycol if (heating1_type == 'Водяний'
                                            and type_ahu in ['Припливна','Припливно-витяжна']) else None
    ws['F35'] = heating2_percent_glycol if (heating2_type == 'Водяний' and confirm_heating2 == 'Так'
                                            and type_ahu in ['Припливна', 'Припливно-витяжна']) else None

    ws['N37'] = '+' if cooling1_type == 'Водяний' and type_ahu in ['Припливна', 'Припливно-витяжна'] else None
    ws['N38'] = '+' if cooling1_type == 'Фреоновий' and type_ahu in ['Припливна', 'Припливно-витяжна'] else None
    ws['N42'] = '+' if (cooling2_type == 'Водяний' and type_ahu in ['Припливна', 'Припливно-витяжна']
                        and confirm_cooling2 == 'Так') else None
    ws['N43'] = '+' if (cooling2_type == 'Фреоновий' and type_ahu in ['Припливна', 'Припливно-витяжна']
                        and confirm_cooling2 == 'Так') else None

    ws['F38'] = cooling1_param if cooling1_type == 'Водяний' and type_ahu in ['Припливна',
                                                                              'Припливно-витяжна'] else None
    ws['F43'] = cooling2_param if (cooling2_type == 'Водяний'
                                   and type_ahu in ['Припливна', 'Припливно-витяжна']
                                   and confirm_cooling2 == 'Так') else None

    if cooling1_type == 'Водяний' and type_ahu in ['Припливна', 'Припливно-витяжна']:
        ws['F40'] = cooling1_glycol
    elif cooling1_type == 'Фреоновий' and type_ahu in ['Припливна', 'Припливно-витяжна']:
        ws['F40'] = freon1

    if cooling2_type == 'Водяний' and type_ahu in ['Припливна', 'Припливно-витяжна'] and confirm_cooling2 == 'Так':
        ws['F45'] = cooling2_glycol
    elif cooling2_type == 'Фреоновий' and type_ahu in ['Припливна', 'Припливно-витяжна'] and confirm_cooling2 == 'Так':
        ws['F45'] = freon2

    ws['F39'] = cooling1_percent_glycol if (cooling1_type == 'Водяний'
                                    and type_ahu in ['Припливна', 'Припливно-витяжна']) else None
    ws['F44'] = cooling2_percent_glycol if (cooling2_type == 'Водяний'
                                            and type_ahu in ['Припливна', 'Припливно-витяжна']
                                            and confirm_cooling2 == 'Так') else None

    ws['D71'] = notes

    ws['C11'] = contact

    ws['C12'] = phone

    ws['C10'] = organization

    timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
    filename = f"output/{user_id}_{timestamp}_filled.xlsx"
    wb.save(filename)

    return filename

















